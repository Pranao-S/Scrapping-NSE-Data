#importing data from excel to python
import pandas as pd
import os
import selenium
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import openpyxl
from pathlib import Path

wait_imp = 10
excel_path = Path("E:\python\stocks list.xlsx")
wb = openpyxl.load_workbook(excel_path)
ws = wb["Sr"] ## should comment out?

print("Step1 -> Reading data from excel")
s_row = 4
s_list = []
avg_val = []
qnty_list = []
while ws.cell(row=s_row, column=2).value!=None:
    s_name = ws.cell(row=s_row, column=2) #stock name
    val_1 = ws.cell(row = s_row, column= 4).value
    qnty = ws.cell(row = s_row, column= 5).value
    s_list.append(s_name)
    avg_val.append(val_1)
    qnty_list.append(qnty)
    s_row += 1    
print ("Company name available in Database")
[print('    ->',name) for name in s_list]
time.sleep(2)
print ('\n')
# create a webdriver object for chrome-option and configure
CO = webdriver.ChromeOptions()
CO.add_experimental_option('useAutomationExtension', False)
CO.add_argument('--ignore-certificate-errors')
CO.add_argument('--start-maximized')
wd = webdriver.Chrome(r'E:\python\chromedriver.exe',options=CO)

print ("Step 2 --> Opening Finance website\n")
wd.implicitly_wait(wait_imp)
wd.get("https://www.nseindia.com") 
time.sleep(5)
print ("******************************************************************************")
print ("                      Getting Live Stock Value !! Please wait ...\n")
for i in range(len(s_list)):
    src = wd.find_element_by_id ("search_str")
    src.send_keys(s_list[i])
    src.send_keys(Keys.RETURN)
    wd.implicitly_wait(wait_imp)
    s_v = wd.find_element_by_xpath("//*[@id='div_nse_livebox_wrap']/div[1]/div[1]/div/div[2]/span[1]")
    ws.cell(row=4+i, column= 3, value = s_v.text)
    
    diff = (avg_val[i] - float(s_v.text))* qnty_list[i]
    per_diff = (diff/(avg_val[i]*qnty_list[i]))*100
    print ("{:>23} -> CMP {:<7} Current P/L->[{:>8.2f}] %P/L -> {:>6.2f}%".format(s_list[i],s_v.text, diff, per_diff))
print ('\n')
print ("Step 3 --> Writing Latest Price into Excel-sheet ....\n")
time.sleep(1)
wb.save(excel_path)
print ("Step 4 --> Successfully Written  \n")
print ("Step 5 --> Closing browser !\n")
print (" ----------------------- FINISHED !! ------------------------")
time.sleep(1)
wd.close()